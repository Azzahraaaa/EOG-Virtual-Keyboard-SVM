# gui_keyboard_svm_direct.py
# EOG Virtual Keyboard + SVM (.pkl) — dengan kontrol keyboard fisik
# Panah (↑↓←→) untuk geser kursor, Enter pilih, Backspace/Del/Space sesuai fungsi.

import os, time, pickle, threading
from collections import deque
from datetime import datetime

import numpy as np
import tkinter as tk
import matplotlib.pyplot as plt
import matplotlib.animation as animation
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import nidaqmx
from nidaqmx.constants import TerminalConfiguration
from openpyxl import Workbook, load_workbook

# ========= PATH / DEVICE =========
MODEL_PATH   = r"D:\ta sempro final\ta azz data\svm_model.pkl"
RT_XLSX_PATH = r"D:\ta sempro final\ta azz data\keyboard_rt_log.xlsx"
DAQ_DEVICE   = "Dev1"

# ========= AMBANG & PARAMETER (tuned) =========
Ts           = 0.10

# Deadband kecil agar event negatif tidak "terpangkas"
thmin, thplus = -0.30, 0.30

# Trigger SVM (kedua channel negatif kuat & konsisten)
NEG_THRESH   = -1.80
MIN_CONSEC   = 2
PROBA_GATE   = 0.70

# Smoothing & kalibrasi baseline
SMOOTH_WIN   = 5
BASELINE_S   = 2.0

# Quiet gate: paksa "diam" saat sinyal kecil
QUIET_THRESH = 0.60
QUIET_CONSEC = 3

# ========= MUAT MODEL =========
with open(MODEL_PATH, "rb") as f:
    svm_pipeline = pickle.load(f)  # Pipeline: StandardScaler -> SVC(probability=True)

# ========= LOCK / DEBOUNCE =========
locked = False
k = 0
lock_durations = {"lirik atas":0.9, "Kedip":1.2, "lirik kiri":0.8, "lirik kanan":0.8, "lirik bawah":1.0}
def set_lock(mv):
    global locked
    locked = True
    threading.Timer(lock_durations.get(mv, 1.0), lambda: release_lock()).start()
def release_lock():
    global locked
    locked = False

# ========= INISIALISASI DAQ =========
task = nidaqmx.Task()
task.ai_channels.add_ai_voltage_chan(f"{DAQ_DEVICE}/ai0", terminal_config=TerminalConfiguration.RSE)
task.ai_channels.add_ai_voltage_chan(f"{DAQ_DEVICE}/ai1", terminal_config=TerminalConfiguration.RSE)

# ========= BUFFER, BASELINE =========
buf_ch1, buf_ch2 = deque(maxlen=SMOOTH_WIN), deque(maxlen=SMOOTH_WIN)
baseline_ch1 = baseline_ch2 = 0.0
calibrated = False

def readdaq():
    """Baca EOG -> kurangi baseline -> deadband -> smoothing."""
    global calibrated
    vals = list(task.read())  # [raw_ch1, raw_ch2]
    if not calibrated:
        return vals
    ch1 = vals[0] - baseline_ch1
    ch2 = vals[1] - baseline_ch2
    if thmin <= ch1 <= thplus: ch1 = 0.0
    if thmin <= ch2 <= thplus: ch2 = 0.0
    buf_ch1.append(ch1); buf_ch2.append(ch2)
    return [sum(buf_ch1)/len(buf_ch1), sum(buf_ch2)/len(buf_ch2)]

# ========= EXCEL LOG =========
log_lock = threading.Lock()
def init_log_excel(path):
    if not os.path.exists(path):
        wb = Workbook(); ws = wb.active; ws.title = "log"
        ws.append(["timestamp","movement","rt_ms","confidence","ch1","ch2","k_sample"])
        wb.save(path)

def log_to_excel(path, ts, movement, rt_ms, conf, ch1, ch2, k_sample):
    """Simpan HANYA: diam / noise / lirik atas / Kedip."""
    if movement not in {"diam", "noise", "lirik atas", "Kedip"}:
        return
    try:
        with log_lock:
            init_log_excel(path)
            wb = load_workbook(path); ws = wb["log"]
            ws.append([
                ts,
                movement,
                (None if rt_ms is None else round(rt_ms, 2)),
                (None if conf   is None else round(conf, 4)),
                round(float(ch1), 4),
                round(float(ch2), 4),
                k_sample
            ])
            wb.save(path)
    except Exception as e:
        print(f"[EXCEL LOG WARNING] {e}")

# ========= STOPWATCH (Response Time) =========
rt_start = None
def rt_begin():
    global rt_start
    if rt_start is None:
        rt_start = time.perf_counter()
def rt_finish():
    global rt_start
    if rt_start is None:
        return None
    rt_ms = (time.perf_counter() - rt_start) * 1000.0
    rt_start = None
    return rt_ms

# ========= GUI =========
class VirtualKeyboard:
    def __init__(self, root):
        self.root = root
        self.root.title("EOG Virtual Keyboard (SVM — with Physical Keyboard Control)")
        self.root.geometry("1000x820"); self.root.configure(bg="white")

        self.text_var = tk.StringVar()
        tk.Entry(root, textvariable=self.text_var, font=("Arial", 16), width=60).pack(pady=8)

        self.movement_label_var = tk.StringVar(value="Menunggu sinyal...")
        tk.Label(root, textvariable=self.movement_label_var, font=("Arial", 16), fg="blue").pack(pady=2)

        self.rt_last_var = tk.StringVar(value="Resp Time (last): - ms")
        self.rt_time_var = tk.StringVar(value="Last action at: -")
        rt_wrap = tk.Frame(root, bg="white"); rt_wrap.pack(pady=2)
        tk.Label(rt_wrap, textvariable=self.rt_last_var, font=("Arial", 11), bg="white").pack(side="left", padx=6)
        tk.Label(rt_wrap, textvariable=self.rt_time_var,  font=("Arial", 11), bg="white").pack(side="left", padx=6)

        # Keyboard sederhana
        self.keys = [
            ["a","b","c","d","e","","",""],
            ["f","g","h","i","j","k","l",""],
            ["m","n","o","DELETE","p","q","r",""],
            ["s","t","BACK","","ENTER","u","v",""],
            ["w","x","y","SPACE","z","1","2",""],
            ["3","4","5","6","7","8","9",""],
            ["0",",",".","?","!","","",""]
        ]
        self.special_keys={"DELETE","BACK","ENTER","SPACE"}
        self.current_row,self.current_col = 3,3
        kb = tk.Frame(root, bg="white"); kb.pack(pady=10)
        self.buttons = []
        for r,row in enumerate(self.keys):
            br=[]
            for c,key in enumerate(row):
                frame=tk.Frame(kb, relief="raised", borderwidth=2, bg="white")
                frame.grid(row=r, column=c, padx=3, pady=3)
                if key:
                    btn=tk.Button(frame, text=key, font=("Arial",12), width=5, height=2,
                                  bg="lightcoral" if key in self.special_keys else "white",
                                  command=lambda k=key: self.on_key_press(k))
                    btn.pack(); br.append(btn)
                else:
                    br.append(None)
            self.buttons.append(br)
        self.highlight_button()

        # Plot realtime
        self.fig,self.ax = plt.subplots(figsize=(7,2))
        self.ax.set_ylim([-6,6]); self.ax.set_xlim(0,100)
        self.line1,=self.ax.plot(range(100), [0]*100, label="CH1")
        self.line2,=self.ax.plot(range(100), [0]*100, label="CH2")
        self.ax.legend(); self.ax.grid()
        self.canvas = FigureCanvasTkAgg(self.fig, master=root)
        self.canvas.get_tk_widget().pack()
        self.ani = animation.FuncAnimation(self.fig, self.update_plot,
                    fargs=(deque([0]*100, 100), deque([0]*100, 100)),
                    interval=int(Ts*1000), blit=True)

        # === KEYBOARD FISIK: BINDINGS ===
        root.bind("<Up>",    lambda e: self.move_cursor("lirik atas"))
        root.bind("<Down>",  lambda e: self.move_cursor("lirik bawah"))
        root.bind("<Left>",  lambda e: self.move_cursor("lirik kiri"))
        root.bind("<Right>", lambda e: self.move_cursor("lirik kanan"))
        root.bind("<Return>",lambda e: self.select_current_key())
        root.bind("<KP_Enter>", lambda e: self.select_current_key())
        root.bind("<BackSpace>", lambda e: self.on_key_press("BACK"))
        root.bind("<Delete>",    lambda e: self.on_key_press("DELETE"))
        root.bind("<space>",     lambda e: self.on_key_press("SPACE"))
        # Ketik langsung huruf/angka/tanda baca → masuk ke input
        for ch in list("abcdefghijklmnopqrstuvwxyz0123456789,.-?!"):
            root.bind(ch, lambda e, _ch=ch: self.on_key_press(_ch))

    # UI helpers
    def update_plot(self, i, ys1, ys2):
        ch1,ch2 = readdaq()
        ys1.append(ch1); ys2.append(ch2)
        self.line1.set_ydata(ys1); self.line2.set_ydata(ys2)
        return self.line1, self.line2

    def on_key_press(self,key):
        if   key=="DELETE": self.text_var.set("")
        elif key=="BACK":   self.text_var.set(self.text_var.get()[:-1])
        elif key=="SPACE":  self.text_var.set(self.text_var.get()+" ")
        elif key=="ENTER":  print("Teks:", self.text_var.get())
        else:               self.text_var.set(self.text_var.get()+key)

    def highlight_button(self):
        for r,row in enumerate(self.buttons):
            for c,btn in enumerate(row):
                if btn: btn.config(bg="lightcoral" if self.keys[r][c] in self.special_keys else "white")
        btn = self.buttons[self.current_row][self.current_col]
        if btn: btn.config(bg="lightblue")

    def move_cursor(self, dir):
        rows, cols = len(self.keys), len(self.keys[0])
        r,c = self.current_row, self.current_col
        if   dir=="lirik atas":   r=(r-1)%rows
        elif dir=="lirik bawah":  r=(r+1)%rows
        elif dir=="lirik kiri":   c=(c-1)%cols
        elif dir=="lirik kanan":  c=(c+1)%cols
        # lewati sel kosong
        while not self.keys[r][c]:
            if dir in ["lirik kiri","lirik kanan"]: c=(c+1)%cols
            else: r=(r+1)%rows
        self.current_row,self.current_col = r,c
        self.highlight_button()

    def select_current_key(self):
        key = self.keys[self.current_row][self.current_col]
        if key: self.on_key_press(key)

    def update_movement_label(self, mv): self.movement_label_var.set(f"Gerakan: {mv}")
    def update_rt(self, rt_ms):
        if rt_ms is None:
            self.rt_last_var.set("Resp Time (last): - ms")
            self.rt_time_var.set("Last action at: -")
        else:
            self.rt_last_var.set(f"Resp Time (last): {rt_ms:.0f} ms")
            self.rt_time_var.set(f"Last action at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}")

# ========= LOOP DETEKSI =========
def main_loop(vk: VirtualKeyboard):
    global k, locked
    neg_count = 0
    quiet_count = 0

    # stopwatch util lokal
    def start_rt_once():
        if 'rt_started' not in main_loop.__dict__ or not main_loop.rt_started:
            rt_begin(); main_loop.rt_started = True
    def reset_rt_flag():
        main_loop.rt_started = False
    reset_rt_flag()

    while True:
        if locked:
            time.sleep(Ts); continue

        k += 1
        ch1, ch2 = readdaq()

        # QUIET GATE
        if abs(ch1) < QUIET_THRESH and abs(ch2) < QUIET_THRESH:
            quiet_count += 1; neg_count = 0
            if quiet_count >= QUIET_CONSEC:
                movement = "diam"
                vk.update_movement_label(movement); vk.update_rt(None)
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                log_to_excel(RT_XLSX_PATH, ts, movement, None, None, ch1, ch2, k)
                reset_rt_flag()
                time.sleep(Ts); continue
        else:
            quiet_count = 0

        # RULE-BASED (tidak dicatat RT/Excel kecuali noise)
        if   thmin < ch1 < thplus and thmin < ch2 < thplus:
            movement = "noise"
            vk.update_movement_label(movement); vk.update_rt(None)
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            log_to_excel(RT_XLSX_PATH, ts, movement, None, None, ch1, ch2, k)
            reset_rt_flag()

        elif ch1 > thplus and ch2 < thmin:
            movement = "lirik kanan"; vk.move_cursor(movement); set_lock(movement)
            vk.update_movement_label(movement); vk.update_rt(None); reset_rt_flag()

        elif ch1 < thmin and ch2 > thplus:
            movement = "lirik kiri"; vk.move_cursor(movement); set_lock(movement)
            vk.update_movement_label(movement); vk.update_rt(None); reset_rt_flag()

        elif ch1 > 1.5 and ch2 > 1.5:
            movement = "lirik bawah"; vk.move_cursor(movement); set_lock(movement)
            vk.update_movement_label(movement); vk.update_rt(None); reset_rt_flag()

        # SVM trigger (keduanya negatif kuat)
        elif ch1 < NEG_THRESH and ch2 < NEG_THRESH:
            neg_count += 1
            if neg_count == 1:
                start_rt_once()
            if neg_count >= MIN_CONSEC:
                X = np.array([[ch1, ch2]], dtype=float)
                try:
                    proba = svm_pipeline.predict_proba(X)[0]
                    conf = float(np.max(proba))
                except Exception:
                    conf = 1.0

                if conf >= PROBA_GATE:
                    pred = int(svm_pipeline.predict(X)[0])  # 1 = lirik atas, 2 = Kedip
                    if pred == 1:
                        movement = "lirik atas"
                        vk.move_cursor(movement); set_lock(movement)
                    elif pred == 2:
                        movement = "Kedip"
                        vk.select_current_key(); set_lock(movement)
                    else:
                        movement = "noise"

                    rt_ms = rt_finish(); reset_rt_flag()
                    vk.update_movement_label(movement); vk.update_rt(rt_ms)
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                    log_to_excel(RT_XLSX_PATH, ts, movement, rt_ms, conf, ch1, ch2, k)
                else:
                    movement = "noise"
                    vk.update_movement_label(movement); vk.update_rt(None)
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                    log_to_excel(RT_XLSX_PATH, ts, movement, None, conf, ch1, ch2, k)
                neg_count = 0
        else:
            # keluar dari kandidat negatif → reset
            neg_count = 0
            movement = "noise"
            vk.update_movement_label(movement); vk.update_rt(None)
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            log_to_excel(RT_XLSX_PATH, ts, movement, None, None, ch1, ch2, k)
            reset_rt_flag()

        time.sleep(Ts)

# ========= START =========
root = tk.Tk()
vk = VirtualKeyboard(root)

# Kalibrasi baseline (diam ±2 dtk)
print("[CALIB] Diamkan mata ~2 detik...")
samples=[]; t0=time.time()
while time.time()-t0 < BASELINE_S:
    samples.append(list(task.read())); time.sleep(Ts)
if samples:
    arr = np.array(samples)
    baseline_ch1 = float(np.median(arr[:,0]))
    baseline_ch2 = float(np.median(arr[:,1]))
calibrated = True
print(f"[CALIB] Baseline CH1={baseline_ch1:.3f}  CH2={baseline_ch2:.3f}")

threading.Thread(target=main_loop, args=(vk,), daemon=True).start()

def on_close():
    try: task.close()
    except: pass
    root.destroy()
root.protocol("WM_DELETE_WINDOW", on_close)
root.mainloop()
